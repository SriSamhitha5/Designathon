import json
import pickle

import streamlit
import streamlit as st
import pandas as pd
import math
import openpyxl
import openai

# Sample user data
# Function to save feedback as JSON in Excel
employee_df = pd.read_excel('employee_data.xlsx')
hr_df = pd.read_excel('HR_DATA.xlsx')

# Function to get reportees for a given manager ID
def get_reportees(manager_id):
    reportees = employee_df[employee_df['ManagerID'] == manager_id]
    return reportees

# Function to save manager rating and comments
def save_manager_rating(employee_id, manager_rating, manager_comments):
    workbook = openpyxl.load_workbook('HR_DATA.xlsx')
    sheet = workbook.active

    # Ensure the EmployeeId column is of the correct type
    hr_df['EmployeeId'] = hr_df['EmployeeId'].astype(str)
    employee_id = str(employee_id)

    # Find the row index of the employee
    row_idx = hr_df.index[hr_df['EmployeeId'] == employee_id].tolist()[0] + 2  # +2 to match Excel row index (1-based)

    # Check if the ManagerRating cell is empty
    if not sheet.cell(row=row_idx, column=3).value:  # Assuming ManagerRating is in the 3rd column
        sheet.cell(row=row_idx, column=3).value = manager_rating
        sheet.cell(row=row_idx, column=5).value = manager_comments  # Assuming ManagerComments is in the 6th column
        workbook.save('HR_DATA.xlsx')
        return True
    return False
def save_feedback(employee_id, feedback_employee, feedback):
    # Load the data
    df = pd.read_excel('HR_DATA.xlsx')
    feedback_data = {"employee_id": employee_id, "feedback": feedback}

    # Ensure the EmployeeId column is of the correct type
    df['EmployeeId'] = df['EmployeeId'].astype(str)
    feedback_employee = str(feedback_employee)

    # Check if the feedback_employee exists in the DataFrame
    if feedback_employee not in df['EmployeeId'].values:
        st.error(f'Employee ID {feedback_employee} not found in the data.')
        return

    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook('HR_DATA.xlsx')
    sheet = workbook.active

    # Find the row index of the feedback_employee
    row_idx = df.index[df['EmployeeId'] == feedback_employee].tolist()[0] + 2  # +2 to match Excel row index (1-based)

    # Get the current feedback cell value (column G, assuming the feedback is stored in the 7th column)
    cell_value = sheet.cell(row=row_idx, column=7).value

    if cell_value:
        # If feedback already exists, load the existing JSON and append the new feedback
        feedback_list = json.loads(cell_value)
        feedback_list.append(feedback_data)
    else:
        # Otherwise, start with a new list
        feedback_list = [feedback_data]

    # Save the updated feedback list back to the cell
    sheet.cell(row=row_idx, column=7).value = json.dumps(feedback_list)

    # Save the workbook
    workbook.save('HR_DATA.xlsx')
def load_user_data(file_path, employee_id):
    df = pd.read_excel(file_path)

    # Convert the DataFrame to a dictionary for easier manipulation
    user_data = {row['EmployeeId']: row.to_dict() for _, row in df.iterrows()}
    # Return data for the specified employee_id
    return user_data[int(employee_id)]

def generate_response(input_text):

    openai.api_type = "azure"
    openai.api_base = "https://testingkey.openai.azure.com/"
    openai.api_version = "2023-09-15-preview"
    openai.api_key = "35f96eb3ffc04868981139be478f99fa"

    response = openai.Completion.create(
        engine="TestingChatModel",
        prompt=input_text,
        temperature=0.2,
        max_tokens=500,
        top_p=1,
        frequency_penalty=0,
        presence_penalty=0,
        best_of=1,
        stop=None)
    return response['choices'][0]['text']
def save_user_data(file_path, data):
    df = pd.DataFrame.from_dict(data, orient='index')
    df.to_excel(file_path, index_label='EmployeeId')

def saveinhr(employeeid,columnname,value):
    file_path = 'HR_DATA.xlsx'
    sheet_name = 'Sheet1'  # Change this to your actual sheet name
    employee_id_to_update = employeeid  # Replace with the actual employee ID
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    # Locate the employee and update the rating
    df.loc[df['EmployeeId'] == employee_id_to_update, columnname] = value

    # Save the changes back to the Excel file
    df.to_excel(file_path, sheet_name=sheet_name, index=False)

# Function to authenticate users
def authenticate(EmployeeId, password):
    user_data = load_user_data("employee_data.xlsx",EmployeeId)
    if user_data["Password"]==int(password):
        return user_data
    else:

        return None


# Main function
def main():
    # Check if user is logged in
    if 'user' not in st.session_state:
        st.session_state.user = None

    st.title('Performance Management System')

    # Login section
    if st.session_state.user is None:
        st.subheader('Login')
        EmployeeId = st.text_input('EmployeeId')
        password = st.text_input('Password', type='password')
        if st.button('Login'):
            user = authenticate(EmployeeId, password)
            if user is not None:
                st.session_state.user = user
                st.experimental_rerun()  # Force UI rerender
            else:
                st.error('Invalid EmployeeId or password. Please try again.')
    else:
        st.subheader(f'Welcome, {st.session_state.user["Name"]} ')
      #  st.write('You have the following options:')
        if st.session_state.user["IsManager"] == 1:
            st.sidebar.header(f'Welcome, {st.session_state.user["Name"]}')
            page = st.sidebar.radio('Go to', ['Submit Self-Ratings', 'View Manager Feedback', 'Set Goals for Next Year',
                                              'Give 360 degree Feedback','Provide Ratings for my reportees'])

            if page == 'Submit Self-Ratings':
                user_data = load_user_data("HR_DATA.xlsx", st.session_state.user["EmployeeId"])
                if math.isnan(user_data["SelfRating"]):
                    st.header('Submit Self-Ratings')
                    rating = st.slider('Your Self Rating (1-5)', 1, 5, 3)
                    if st.button('Submit Rating'):
                        saveinhr(st.session_state.user["EmployeeId"], 'SelfRating', rating)
                        st.session_state.user['SelfRating'] = rating

                        st.success('Self-rating submitted!')
                else:
                    st.header('You have already Submitted self ratings for approval')
                    st.write('rating submitted by you:' + str(user_data["SelfRating"]))

            elif page == 'View Manager Feedback':
                user_data = load_user_data("HR_DATA.xlsx", st.session_state.user["EmployeeId"])
                if math.isnan(user_data["ManagerRating"]):
                    st.header('manager review not completed yet')
                else:
                    st.header('View Manager Feedback')
                    manager_feedback = user_data["ManagerComments"]
                    st.write(user_data["ManagerRating"])
                    st.text_area('Manager Feedback', manager_feedback, height=100, disabled=True)

            elif page == 'Set Goals for Next Year':
                st.header('Set Goals for Next Year')
                user_data = load_user_data("HR_DATA.xlsx", st.session_state.user["EmployeeId"])
                # goals = st.text_area('Your Goals', st.session_state.user_info.get('Goals', ''))
                if pd.isna(user_data["GoalsForNextYear"]):
                    st.write('Goals for next year not completed yet')
                    plan = st.text_area('Goals Planned for next year')
                    if st.button('Submit Rating'):
                        saveinhr(st.session_state.user["EmployeeId"], 'GoalsForNextYear', plan)
                        st.success('Self-rating submitted!')

                else:
                    st.header('Goals set for Next Year')
                    st.write(user_data["GoalsForNextYear"])


            elif page == 'Give 360 degree Feedback':
                df = pd.read_excel('employee_data.xlsx')
                st.header('Give Feedback about Manager and Teammates')
                manager_id = df.loc[df['EmployeeId'] == st.session_state.user["EmployeeId"], 'ManagerID'].values[0]
                employees_with_same_manager = df[
                    (df['ManagerID'] == manager_id) & (df['EmployeeId'] != st.session_state.user["EmployeeId"])]
                if not employees_with_same_manager.empty:
                    options = employees_with_same_manager.apply(lambda x: f"{x['Name']} ({x['EmployeeId']})",
                                                                axis=1).tolist()

                    selected_option = st.selectbox('Select an employee to give feedback about:', options)

                    if selected_option:
                        # Extract the name and employee ID from the selected option
                        selected_name = selected_option.split(' (')[0]
                        selected_employee_id = selected_option.split(' (')[1][:-1]

                        st.write(f"Selected Employee Name: {selected_name}")
                        st.write(f"Selected Employee ID: {selected_employee_id}")

                        feedback = st.text_area("Enter your feedback")

                if st.button('Submit Feedback'):
                    save_feedback(st.session_state.user["EmployeeId"], selected_employee_id, feedback)
                    st.write('Feedback submitted successfully!')
            elif page == 'Provide Ratings for my reportees':
                reportees = get_reportees(st.session_state.user["EmployeeId"])
                st.write('Your Reportees:')
                # Create a dictionary to map names to employee IDs
                name_to_id = dict(zip(reportees['Name'], reportees['EmployeeId']))

                selected_employee_name = st.selectbox(
                    'Select an employee to view details:',
                    reportees['Name']
                )

                # Map the selected name back to the employee ID
                selected_employee_id = name_to_id[selected_employee_name]

                if selected_employee_id:
                    employee_details = hr_df[hr_df['EmployeeId'] == selected_employee_id]
                    st.write('Employee Details:')

                    detail_columns = [
                        'SelfRating', 'GoalsForNextYear', '360Feedback', 'IsBillable',
                        'Avg_WFO_Days', 'Company_Initiatives', 'LeraningActivities Completed',
                        'Avg_Time_Office'
                    ]
                    for column in detail_columns:
                        st.write(f"**{column.replace('_', ' ')}:** {employee_details[column].values[0]}")

                    # Check if ManagerRating and ManagerComments already exist
                    manager_rating = employee_details['ManagerRating'].values[0]
                    manager_comments = employee_details['ManagerComments'].values[0]

                    if pd.notna(manager_rating) and pd.notna(manager_comments):
                        st.write(f"**Existing Manager Rating:** {manager_rating}")
                        st.write(f"**Existing Manager Comments:** {manager_comments}")
                    else:

                        manager_rating = st.slider('Rate the employee (1-5):', 1, 5)
                        manager_comments = st.text_area('Manager Comments:')
                        input_data = {
                            'is_billed': [employee_details['IsBillable'].values[0]],
                            # Example input values, you can replace them with your own data
                            'avg_working_hours': [employee_details['Avg_Time_Office'].values[0]],
                            'learning_activities_completed': [
                                employee_details['LeraningActivities Completed'].values[0]],
                            'days_worked_last_year': [employee_details['Avg_WFO_Days'].values[0]],
                            'company_initiatives_taken': [employee_details['Company_Initiatives'].values[0]],
                            'employee_self_rating': [employee_details['SelfRating'].values[0]]
                        }



                        comment_visible = False

                        # Button to toggle visibility
                        if st.button("Show Udemy courses suggested from 360 feedback"):
                            # Toggle the visibility of the comment
                            comment_visible = not comment_visible

                        # Display the comment if visibility is True
                        if comment_visible:
                            suggested_courses = generate_response(
                                'configure yourself for assisting me in employee rating i have an employee who got comments from peers like  ' +
                                str(employee_details['360Feedback'].values[0]) + ' suggest some udemy courses for him')
                            saveinhr(selected_employee_id, 'SuggestedCourses', suggested_courses)
                            st.write(suggested_courses)
                        comment_visible2 = False

                        # Button to toggle visibility
                        if st.button("Show personalized development plan for employee"):
                            # Toggle the visibility of the comment
                            comment_visible2 = not comment_visible2

                        # Display the comment if visibility is True
                        if comment_visible2:
                            Development_plan = generate_response(
                                'configure yourself for assisting me in employee rating i have an employee who got comments from peers like' +
                                str(employee_details['360Feedback'].values[
                                        0]) + 'give a personalized plan for his development and additional detail :' + str(
                                    input_data))
                            saveinhr(selected_employee_id, 'Development_Plan', Development_plan)
                            st.write(Development_plan)
                        comment_visible3 = False

                        # Button to toggle visibility
                        if st.button("Show OpenAI comment about employee"):
                            # Toggle the visibility of the comment
                            comment_visible3 = not comment_visible3

                        # Display the comment if visibility is True
                        if comment_visible3:
                            AI_COMMENT = generate_response(
                                'configure yourself for assisting me in employee rating i have an employee who got comments from peers like ' +
                                str(employee_details['360Feedback'].values[
                                        0]) + ' give a personalized plan for his development in the tone of saying from my perspective what yor are thinking about him and additional detail :' + str(
                                    input_data))
                            saveinhr(selected_employee_id, 'AI_Comment', AI_COMMENT)
                            st.write(AI_COMMENT)
                        if st.button('Assist Rating this Employee:'):


                            with open('manager_rating_model.pkl', 'rb') as f:
                                model = pickle.load(f)
                            input_df = pd.DataFrame(input_data)

                            # Use the loaded model to make predictions on the input data
                            predicted_manager_rating = model.predict(input_df)
                            st.write('Rating predicted by model: '+ str(predicted_manager_rating))

                        if st.button('Submit Rating and Comments'):
                            save_manager_rating(selected_employee_id, manager_rating, manager_comments)
                            st.write('Rating and comments submitted successfully!')

        else:
            st.sidebar.header(f'Welcome, {st.session_state.user["Name"]}')
            page = st.sidebar.radio('Go to', ['Submit Self-Ratings', 'View Manager Feedback', 'Set Goals for Next Year',
                                              'Give 360 degree Feedback'])

            if page == 'Submit Self-Ratings':
                user_data = load_user_data("HR_DATA.xlsx", st.session_state.user["EmployeeId"])
                if math.isnan(user_data["SelfRating"]):
                    st.header('Submit Self-Ratings')
                    rating = st.slider('Your Self Rating (1-5)', 1, 5, 3)
                    if st.button('Submit Rating'):
                        saveinhr(st.session_state.user["EmployeeId"],'SelfRating',rating)
                        st.session_state.user['SelfRating'] = rating

                        st.success('Self-rating submitted!')
                else:
                    st.header('You have already Submitted self ratings for approval')
                    st.write('rating submitted by you:'+str(user_data["SelfRating"]))

            elif page == 'View Manager Feedback':
                user_data = load_user_data("HR_DATA.xlsx", st.session_state.user["EmployeeId"])
                if math.isnan(user_data["ManagerRating"]):
                    st.header('manager review not completed yet')
                else:
                    st.header('View Manager Feedback')
                    manager_feedback = user_data["ManagerComments"]
                    st.write(user_data["ManagerRating"])
                    st.text_area('Manager Feedback', manager_feedback, height=100, disabled=True)

                    st.text_area('Suggested Courses', user_data["SuggestedCourses"], height=100, disabled=True)

                    st.text_area('Development_Plan', user_data["Development_Plan"], height=100, disabled=True)

                    st.text_area('AI_Comment', user_data["AI_Comment"], height=100, disabled=True)


            elif page == 'Set Goals for Next Year':
                st.header('Set Goals for Next Year')
                user_data = load_user_data("HR_DATA.xlsx", st.session_state.user["EmployeeId"])
                # goals = st.text_area('Your Goals', st.session_state.user_info.get('Goals', ''))
                if pd.isna(user_data["GoalsForNextYear"]):
                    st.write('Goals for next year not completed yet')
                    plan = st.text_area('Goals Planned for next year')
                    if st.button('Submit Rating'):
                        saveinhr(st.session_state.user["EmployeeId"], 'GoalsForNextYear', plan)
                        st.success('Self-rating submitted!')

                else:
                    st.header('Goals set for Next Year')
                    st.write(user_data["GoalsForNextYear"])


            elif page == 'Give 360 degree Feedback':
                df = pd.read_excel('employee_data.xlsx')
                st.header('Give Feedback about Manager and Teammates')
                manager_id = df.loc[df['EmployeeId'] == st.session_state.user["EmployeeId"], 'ManagerID'].values[0]
                employees_with_same_manager = df[(df['ManagerID'] == manager_id) & (df['EmployeeId'] != st.session_state.user["EmployeeId"])]
                if not employees_with_same_manager.empty:
                    options = employees_with_same_manager.apply(lambda x: f"{x['Name']} ({x['EmployeeId']})",
                                                                axis=1).tolist()

                    selected_option = st.selectbox('Select an employee to give feedback about:', options)

                    if selected_option:
                        # Extract the name and employee ID from the selected option
                        selected_name = selected_option.split(' (')[0]
                        selected_employee_id = selected_option.split(' (')[1][:-1]

                        st.write(f"Selected Employee Name: {selected_name}")
                        st.write(f"Selected Employee ID: {selected_employee_id}")

                        feedback = st.text_area("Enter your feedback")

                if st.button('Submit Feedback'):
                        save_feedback(st.session_state.user["EmployeeId"], selected_employee_id, feedback)
                        st.write('Feedback submitted successfully!')



        # Logout button
        if st.button('Logout'):
            st.session_state.user = None
            st.info('Logged out successfully.')
            st.experimental_rerun()  # Force UI rerender


if __name__ == '__main__':
    main()
